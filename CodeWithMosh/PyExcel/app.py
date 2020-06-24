# work with excel spreadsheets
# project PyExcel
# in project we have transcations.xlsx(remove transactions2.xlsx if doing again)
# pipenv install openpyxl
# pipenv shell to activate virtual env
# cd C:\Users\Gopi\.virtualenvs\PyExcel-xHoxv4gv\Scripts
# then activate if going from outside
# select virtual environment from status bar of vs code
import openpyxl

# wb=openpyxl.workbook() #to create new workbook
# to load an existing workbook
wb = openpyxl.load_workbook("transactions.xlsx")
# has three columns, transaction_id,product_id,price and has sample data, it can be created on our own
# we have attribute called sheetnames
print(wb.sheetnames)  # Sheet1
sheet = wb["Sheet1"]  # create a sheet object
# we can create a new sheet or delete an existing one
# wb.create_sheet("Sheet2",0) # index 0 puts it before Sheet1
# wb.remove_sheet("Sheet2")
# accessing the contents
cell = sheet["a1"]  # a1 is the position of the cell,
# cell object has different attributes, cell.row,cell.value,cellcolumn,cell.coordinate etc

# one more way to access
# cell=sheet.cell(row=1,column=1)
print(sheet.max_row)
print(sheet.max_column)
for row in range(1, sheet.max_row+1):  # first row is 1 , hence 1
    for column in range(1, sheet.max_column+1):
        cell = sheet.cell(row, column)
        print(cell.value)

# we can access a range of cells
column = sheet["a"]  # get  cells in a column
# returns a tuple
print(column)  # we get tuple of cell objects

# we can also work with a range of columns, all cells in column a to c
cells = sheet["a":"c"]
print(cells)  # we get tuple of tuples

# we can slo use coordinates
cells2 = sheet["a1":"c3"]
print(cells2)
# for rows
sheet[1:3]

# to append
sheet.append([1, 2, 3])
# ton insert rows,insert cols, dlete rows,delete cols
# sheet.insert_rows
wb.save("transactions2.xlsx")  # save changes to new excel


# command query violation in app1.py which is the next section
